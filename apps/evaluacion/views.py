#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Código desarrollado por Daniel Mansilla
from django.shortcuts import render, redirect
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, View, DetailView

from django.core.urlresolvers import reverse, reverse_lazy
from django.utils import timezone
from django.core import serializers
from django.http.response import HttpResponse
from django.http import JsonResponse
import json

from django.contrib.auth.models import User

from Proyecto.utilities import render_pdf, generate_pdf

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, cm
import random


from .models import *
from .forms import *
# from apps.preguntas.forms import PreguntaForm


def evaluacion_list(request):
	usuario = User.objects.get(id=request.user.id)
	tipo = usuario.perfilusuario.tipo_usuario
	evaluaciones = Evaluacion.objects.filter(docente=usuario)

	context = {
		'usuario': usuario,
		'tipo': tipo,
		'evaluaciones': evaluaciones,
	}

	return render(request, 'apps/evaluaciones/evaluacionList.html', context)


def evaluacion_step1(request):
	if request.POST:
		form = EvaluacionForm(request.POST)
		if form.is_valid():
			form = form.save(commit=False)
			form.docente = request.user
			form.save()

		return redirect('evaluaciones:evaluacion_step2', form.pk)
	else:
		form = EvaluacionForm()


	context = {
		'form': form,
	}


	return render(request, 'apps/evaluaciones/evaluacion_step1.html', context)

def evaluacion_step1_back(request, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)

	if request.method == 'GET':
		form = EvaluacionForm(instance=evaluacion)
	else:
		form = EvaluacionForm(request.POST, instance=evaluacion)
		
		if form.is_valid():
			form = form.save(commit=False)
			form.docente = request.user
			form.save()
		# evaluacion.preguntas.add(*pregunta)


		return redirect('evaluaciones:evaluacion_step2', evaluacion.pk)
	# if request.method == 'GET':
	# 	form = EvaluacionForm(instance=evaluacion)
	# else:
	# 	form = EvaluacionForm(instance=evaluacion)

	context = {
		'evaluacion': evaluacion,
		'form': form,
	}

	return render(request, 'apps/evaluaciones/evaluacion_step1_back.html', context)

def evaluacion_step2(request, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)
	preguntas = Pregunta.objects.filter(asignatura=evaluacion.curso.asignatura.id, status=True).order_by('?')

	if request.method == 'GET':
		form = EvaluacionForm(instance=evaluacion)
	else:
		form = EvaluacionForm(request.POST, instance=evaluacion)
		pregunta = request.POST.getlist('preguntas', None)
		
		if form.is_valid():
			form = form.save(commit=False)

			for p in pregunta:
				try:
					form.preguntas.add(p)

					# ---------------------------
					# Sumale contador a pregunta 
					obj_pregunta =  Pregunta.objects.get(id=int(p))
					obj_pregunta.cant_usada = obj_pregunta.cant_usada + 1
					obj_pregunta.save()
					# ---------------------------

					# contenidos = Unidad.objects.filter()
					# evaluacion.contenidos.add
				except:
					pass
			form = form.save()
		# evaluacion.preguntas.add(*pregunta)


		return redirect('evaluaciones:evaluacion_step3', evaluacion.pk)
	# if request.method == 'GET':
	# 	form = EvaluacionForm(instance=evaluacion)
	# else:
	# 	form = EvaluacionForm(instance=evaluacion)

	context = {
		'evaluacion': evaluacion,
		'preguntas': preguntas,
		'form': form,
	}

	return render(request, 'apps/evaluaciones/evaluacion_step2.html', context)

def evaluacion_step3(request, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)
	# preguntas = Pregunta.objects.filter(asignatura=evaluacion.seccion.asignatura.id, status=True).order_by('?')
	instrucciones = "No se permite el uso de smartphones, tablets u otro dispositivo tecnológico.\nUse lápiz pasta para responder las preguntas.\nEl uso de lapiz grafito no tendrá derecho a reclamos."

	if request.method == 'GET':
		form = EvaluacionForm(instance=evaluacion)
	else:
		form = EvaluacionForm(request.POST, instance=evaluacion)
		print request.POST
		if form.is_valid():
			form.save()
		return redirect('evaluaciones:evaluacion_step4', evaluacion.pk)

	context = {
		'evaluacion': evaluacion,
		'form': form,
		'instrucciones': instrucciones,
	}

	return render(request, 'apps/evaluaciones/evaluacion_step3.html', context)

def evaluacion_step4(request, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)
	preguntas = evaluacion.preguntas.all()
	instrucciones = "No se permite el uso de smartphones, tablets u otro dispositivo tecnológico.\nUse lápiz pasta para responder las preguntas.\nEl uso de lapiz grafito no tendrá derecho a reclamos."

	print preguntas
	if request.method == 'GET':
		form = EvaluacionForm(instance=evaluacion)
	else:
		form = EvaluacionForm(request.POST, instance=evaluacion)
		puntaje = request.POST.getlist('pje_pregunta')
		if form.is_valid():
			form.save(commit=False)
			puntos = 0
			for p in puntaje:
				puntos += int(p)
			evaluacion.puntaje_total = puntos
			form.puntaje_total = evaluacion.puntaje_total
			form.save()

		for p in range(len(preguntas)):
			form2 = PuntajePregunta()
			try:
				puntaje_p = PuntajePregunta.objects.get(pregunta=preguntas[p].id, evaluacion=evaluacion)
				print 'ya existe'
				puntaje_p.pregunta = preguntas[p]
				puntaje_p.puntaje = puntaje[p]
				puntaje_p.save()
			except: 
				form2.evaluacion = evaluacion
				form2.pregunta = preguntas[p]
				form2.puntaje = puntaje[p]
				form2.save()


		return redirect("evaluaciones:evaluacion_detail", evaluacion.pk)


	context = {
		'evaluacion': evaluacion,
		'form': form,
		'instrucciones': instrucciones,
	}

	return render(request, 'apps/evaluaciones/evaluacion_step4.html', context)

def evaluacion_rapida_step1(request):
	if request.POST:
		form = EvaluacionForm(request.POST)

		#capturar la cantidad de preguntas elegidas
		cantidad_desarrollo = request.POST.get('p_desarrollo')
		cantidad_alternativa = request.POST.get('p_alternativas')
		cantidad_pareados = request.POST.get('p_pareados')

		#transformar a enteros
		p1 = int(cantidad_desarrollo)
		p2 = int(cantidad_alternativa)
		p3 = int(cantidad_pareados)

		desarrollo = []
		alternativas = []
		pareados = []

		#capturar asignatura
		asignatura = request.POST.get('asignatura')
		#capturar contenidos
		contenidos = request.POST.getlist('contenidos', None)

		# Queryset para capturar las preguntas
		contents = []
		for c in contenidos:
			contents += [c]


		for i in range(p1): #	preguntas de desarrollo
			p_desarrollo = Pregunta.objects.filter(tipo_pregunta='Pregunta de desarrollo', 
				asignatura=asignatura, status=True, unidad=contents[random.randint(0,len(contents)-1)]).order_by('?')[:1]
			desarrollo += list(p_desarrollo)


		for i in range(p2): #	preguntas de alternativas
			p_alternativas = Pregunta.objects.filter(tipo_pregunta='Seleccion multiple', 
				asignatura=asignatura, status=True, unidad=contents[random.randint(0,len(contents)-1)]).order_by('?')[:1]
			p_alternativas = list(p_alternativas)
			alternativas += p_alternativas

		for i in range(p3): #	preguntas de terminos pareados
			p_pareados = Pregunta.objects.filter(tipo_pregunta='Terminos pareados', 
				asignatura=asignatura, status=True, unidad=contents[random.randint(0,len(contents)-1)]).order_by('?')[:1]
			p_pareados = list(p_pareados)
			pareados += p_pareados


		# desarrollo = Pregunta.objects.filter(tipo_pregunta='Pregunta de desarrollo', 
		# 	asignatura=asignatura, status=True, unidad=contents[random.randint(0,len(contents)-1)]).order_by('?')[:p1]
		# alternativas = Pregunta.objects.filter(tipo_pregunta='Seleccion multiple', 
		# 	asignatura=asignatura, status=True, unidad=contents[random.randint(0,len(contents)-1)]).order_by('?')[:p2]
		# pareados = Pregunta.objects.filter(tipo_pregunta='Terminos pareados', 
		# 	asignatura=asignatura, status=True, unidad=contents[random.randint(0,len(contents)-1)]).order_by('?')[:p3]

		desarrollo = list(desarrollo)
		print desarrollo
		alternativas = list(alternativas)
		pareados = list(pareados)

		#generar lista con todas las preguntas 
		pregunta = desarrollo, alternativas, pareados
		pregunta = list(pregunta)
		print pregunta
		# print pregunta
		# ========================== #
		tipo_evaluacion = request.POST.get('tipo_evaluacion') # get tipo de evaluación
		a = Asignatura.objects.get(id=asignatura)
		titulo = tipo_evaluacion + ' - ' + a.nombre
		titulo = titulo.upper()

		instucciones = "No se permite el uso de smartphones, tablets u otro dispositivo tecnológico.\nUse lápiz pasta para responder las preguntas.\nEl uso de lapiz grafito no tendrá derecho a reclamos."
		puntajes = request.POST.get('pje_automatico')


		nombre_usuario = "%s %s %s" %(request.user.perfilusuario.nombres, request.user.perfilusuario.apellido1, request.user.perfilusuario.apellido2)
		if form.is_valid():
			form = form.save(commit=False)
			form.docente = request.user
			form.titulo = titulo
			form.instrucciones = instucciones
			form.datos = nombre_usuario
			# if not puntajes:
			# 	form.puntaje_total = 100
			form.save()

			for p in pregunta:
				# print "-" * 100
				# print p
				# form.preguntas.add(*p)
				# ---------------------------
				# Sumale contador a pregunta 
				if type(p) == list:
					# print "es lista"
					for i in p:
						print i.pk
						obj_pregunta =  Pregunta.objects.get(id=int(i.pk))
						obj_pregunta.cant_usada = obj_pregunta.cant_usada + 1
						obj_pregunta.save()
						form.preguntas.add(i)
				else:
					print "no es lista"
					obj_pregunta =  Pregunta.objects.get(id=int(p.pk))
					obj_pregunta.cant_usada = obj_pregunta.cant_usada + 1
					obj_pregunta.save()
					form.preguntas.add(p)
				# ---------------------------

		preguntas = form.preguntas.all()



		if not puntajes:
			return redirect("evaluaciones:evaluacion_rapida_step2", form.id)
		else:
			#dejar en True el estado de la evaluación generada
			form.disponible = True
			form.save()

			pje = 0


			for p in range(len(preguntas)):
				form2 = PuntajePregunta()
				try:
					puntaje_p = PuntajePregunta.objects.get(pregunta=preguntas[p].id, evaluacion=evaluacion)
					print 'ya existe'
					puntaje_p.pregunta = preguntas[p]
					puntaje_p.puntaje = puntaje
					puntaje_p.save()
				except: 
					form2.evaluacion = form
					form2.pregunta = preguntas[p]
					#definir puntaje de pregunta según dificultad
					if form2.pregunta.dificultad == 'Basica':
						puntaje = random.randint(10, 20)
					elif form2.pregunta.dificultad == 'Intermedia':
						puntaje = random.randint(21, 40)
					elif form2.pregunta.dificultad == 'Avanzada':
						puntaje = random.randint(41, 60)
					else:
						puntaje = random.randint(10, 20)
						

					print puntaje
					form2.puntaje = puntaje
					form2.save()

					pje += puntaje
			form.puntaje_total = pje 
			form.save()

			return redirect("evaluaciones:evaluacion_detail", form.id)


	context = {

	}
	return render(request, 'apps/evaluaciones/evaluacion_rapida_step1.html', context)

def evaluacion_rapida_step2(request, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)
	preguntas = evaluacion.preguntas.all()
	print preguntas
	if request.method == 'GET':
		form = EvaluacionForm(instance=evaluacion)
	else:
		form = EvaluacionForm(request.POST, instance=evaluacion)
		puntaje = request.POST.getlist('pje_pregunta')
		if form.is_valid():
			form.save(commit=False)
			puntos = 0
			for p in puntaje:
				puntos += int(p)
			evaluacion.puntaje_total = puntos
			form.puntaje_total = evaluacion.puntaje_total
			form.save()

		for p in range(len(preguntas)):
			form2 = PuntajePregunta()
			try:
				puntaje_p = PuntajePregunta.objects.get(pregunta=preguntas[p].id, evaluacion=evaluacion)
				print 'ya existe'
				puntaje_p.pregunta = preguntas[p]
				puntaje_p.puntaje = puntaje[p]
				puntaje_p.save()
			except: 
				form2.evaluacion = evaluacion
				form2.pregunta = preguntas[p]
				form2.puntaje = puntaje[p]
				form2.save()


		return redirect("evaluaciones:evaluacion_detail", evaluacion.pk)


	context = {
		'evaluacion': evaluacion,
		'form': form,
	}

	return render(request, 'apps/evaluaciones/evaluacion_rapida_step2.html', context)

def evaluacion_detail(request, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)
	preguntas = evaluacion.preguntas.all()
	puntaje = PuntajePregunta.objects.filter(evaluacion=evaluacion.id)

	context = {
		'evaluacion': evaluacion,
		'preguntas': preguntas,
		'puntaje': puntaje,
	}

	return render(request, 'apps/evaluaciones/evaluacion_detail.html', context)

#	Excel
def calificaciones(request, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)
	curso = evaluacion.curso
	estudiantes = curso.estudiantes.all().order_by('apellido1')
	preguntas = evaluacion.preguntas.all()

	#filas y columnas
	fila = 2
	columna = 1

	#generar excel
	wb = Workbook()
	# ws = wb.create_sheet("Calificaciones",0)
	ws = wb.active
	ws.title = 'Calificaciones'
	# ws.font = ws.font.copy(bold=True, italic=True)
	# Cabeceras
	# ws.cell(row=1, column=columna).value = 'Nombre Estudiante'
	# ws.cell(row=1, column=columna+1).value = 'RUN'
	a1 = ws.cell(row=1, column=columna, value='Nombre Estudiante')
	a2 = ws.cell(row=1, column=columna+1, value='RUN')
	a3 = ws.cell(row=1, column=columna+2, value='email')
	a1.font = Font(bold=True)
	a2.font = Font(bold=True)
	a3.font = Font(bold=True)

	columna = 4
	#Cabecera preguntas
	count = 1
	for celda in preguntas:
		puntaje = PuntajePregunta.objects.get(evaluacion=evaluacion, pregunta=celda)
		a4 = ws.cell(row=1, column=columna, value = "Pregunta %s \n(Puntaje total: %s)" % (count,puntaje.puntaje))
		a4.font = Font(bold=True)

		# a5 = ws.cell(row=2, column=columna, value = "")
		# a5.font = Font(bold=True)


		columna += 1
		count += 1
	ws.cell(row=1, column=columna).value = ''

	# Nombre y rut de estudiantes
	for e in estudiantes:
		nombre = "%s %s, %s %s" % (e.apellido1, e.apellido2, e.nombre1, e.nombre2)
		print nombre
		b1 = ws.cell(row=fila, column=1, value = nombre)
		b2 = ws.cell(row=fila, column=2, value = e.rut)
		b3 = ws.cell(row=fila, column=3, value = e.email)
		b1.font = Font(italic=True)
		b2.font = Font(italic=True)
		b3.font = Font(italic=True)
		fila += 1




	nombre_archivo = 'Calificaciones_%s_%s.xlsx' % (evaluacion.curso.nrc ,evaluacion.curso.asignatura.codigo)
	response = HttpResponse(content_type="application/ms-excel")
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response

def get_calificaciones(request, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)

	if request.POST:
		excel = request.FILES['notas'].read() # capturar archivo
		wb = openpyxl.load_workbook(filename=BytesIO(excel)) # iniciar archivo
		hojas = wb.get_sheet_names() # capturar nombre de hojas del archivo
		hoja = wb.get_sheet_by_name(hojas[0]) #utilizar la primera hoja del documento

		total_filas = hoja.max_row # capturar valor maximo de filas a leer
		total_columnas = hoja.max_column # capturar valor maximo de columnas
		print total_columnas

		#obtener puntaje de preguntas
		puntaje_alumno = []
		for i in range(2, total_filas+1):
			pje = 0.0
			for j in range(4, total_columnas+1):
				puntaje_invidual = hoja.cell(row=i,column=j).value
				print puntaje_invidual
				if puntaje_invidual is None:
					puntaje_invidual = 0
				pje = pje + float(puntaje_invidual)
			# pje = str(pje)
			puntaje_alumno += [pje]

		#obtener estudiantes
		estudiantes = []
		for i in range(2, total_filas+1):
			rut = hoja.cell(row=i,column=2).value
			print rut
			estudiantes += [evaluacion.curso.estudiantes.get(rut=rut)]
		print estudiantes

		exigencia = request.POST.get('exigencia')
		exigencia = float(exigencia)
		print evaluacion.puntaje_total
		pje_maximo = float(evaluacion.puntaje_total)
		nota_minima = request.POST.get('nota_minima')
		nota_minima = float(nota_minima)
		nota_maxima = 7.0
		nota_aprobacion = 4.0

		pje_aprobacion = (exigencia*pje_maximo)/100.0

		calificacion_promedio = []
		get_calificacion = []

		j = 0			
		for i in range(len(puntaje_alumno)):
			form = CalificacionForm()
			# if form.is_valid():
			form = form.save(commit=False)
			form.evaluacion = evaluacion 
			form.puntaje_obtenido = float(puntaje_alumno[i])

			#Calculo de calificacion segun puntaje del alumno
			if float(puntaje_alumno[i]) <= pje_aprobacion:
				nota = (((nota_aprobacion-nota_minima)/(pje_aprobacion-0.0))*(float(puntaje_alumno[i])-0.0))+nota_minima
			else:
				nota = (((nota_maxima-nota_aprobacion)/(pje_maximo-pje_aprobacion))*(float(puntaje_alumno[i])-pje_aprobacion))+nota_aprobacion

			calificacion_promedio += [nota]

			# try:
			# 	calificacionObj = Calificacion.objects.get

			#	GUARDAR CALIFICACION
			form.calificacion = round(nota,2)
			form.status = True
			form.save()

			get_calificacion += [form.id]
			print get_calificacion
			

			# CALIFICACION POR ALUMNO
			form3 = EstudianteCalificacionForm()
			form3 = form3.save(commit=False)
			form3.evaluacion = evaluacion
			form3.calificacion = form
			get_estudiante = Estudiante.objects.get(rut=estudiantes[j])
			form3.estudiante = get_estudiante
			print "-" * 100
			form3.save()
			# estudiantes[j].evaluaciones.add(evaluacion)
			# estudiantes[j].calificaciones.add(form)
			j+=1

		promedio = 0.0
		for prom in calificacion_promedio:
			promedio += prom

		promedio = promedio/len(calificacion_promedio)

		pje_promedio = 0.0
		for pje_prom in puntaje_alumno:
			pje_promedio += pje_prom

		pje_promedio = pje_promedio/len(puntaje_alumno)

		print promedio 

		

		form2 = PromedioForm()
		# print form2

		try: 
			promedio_prom = Promedio.objects.get(evaluacion=evaluacion)
			promedio_prom.promedio = promedio
			promedio_prom.pje_promedio = pje_promedio
			promedio_prom.save()
			for c in get_calificacion:
				promedio_prom.calificaciones.add(c)
				
		except:

			form2 = form2.save(commit=False)
			form2.evaluacion = evaluacion
			form2.curso = evaluacion.curso
			form2.promedio = promedio
			form2.pje_promedio = pje_promedio
			form2.save()
			form2.calificaciones.add(*get_calificacion)

		# print get_calificacion
		print "save"


				
		return redirect('evaluaciones:listar_evaluacion')

	else:
		form = CalificacionForm()
		

	context = {
		'evaluacion': evaluacion,
		'form': form,
	}

	return render(request, 'apps/evaluaciones/calificacion_agregar.html' ,context)

#	PDF
# def evaluacion_pdf(request, evaluacion_id, *args, **kwargs):
# 	evaluacion = Evaluacion.objects.get(id=evaluacion_id)
# 	preguntas = evaluacion.preguntas.all()
# 	puntaje = PuntajePregunta.objects.filter(evaluacion=evaluacion.id)

# 	context = {
# 		'evaluacion': evaluacion,
# 		'preguntas': preguntas,
# 		'puntaje': puntaje,
# 	}
# 	pdf = render_pdf('apps/evaluaciones/pdf/evaluacion.html', context)
# 	return HttpResponse(pdf, content_type="application/pdf")

def solucion_pdf(response, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)
	preguntas = evaluacion.preguntas.all()
	puntaje = PuntajePregunta.objects.filter(evaluacion=evaluacion.id)

	context = {
		'evaluacion': evaluacion,
		'preguntas': preguntas,
		'puntaje': puntaje,
	}

	resp = HttpResponse(content_type='application/pdf')
	result = generate_pdf('apps/evaluaciones/pdf/solucion.html', context , file_object=resp)
	return result


def evaluacion_pdf(response, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)
	preguntas = evaluacion.preguntas.all()
	puntaje = PuntajePregunta.objects.filter(evaluacion=evaluacion.id)

	context = {
		'evaluacion': evaluacion,
		'preguntas': preguntas,
		'puntaje': puntaje,
	}
	resp = HttpResponse(content_type='application/pdf')
	result = generate_pdf('apps/evaluaciones/pdf/evaluacion.html', context , file_object=resp)
	return result


#	Calificacion Estudiantes
def estudiante_calificacion(request, evaluacion_id):
	evaluacion = Evaluacion.objects.get(id=evaluacion_id)
	estudiantes = EstudianteCalificacion.objects.filter(evaluacion=evaluacion)

	context = {
		'evaluacion': evaluacion,
		'estudiantes': estudiantes,
	}

	return render(request, 'apps/evaluaciones/notas_alumnos.html', context)

# AJAX
class ContenidoFiltroAjax(TemplateView):
	def get(self, request, *args, **kwargs):
		asignatura = request.GET['asignatura']
		asignatura = int(asignatura)
		contenidos = Unidad.objects.filter(asignatura=asignatura)
		data = serializers.serialize('json', contenidos)
		return HttpResponse(data, content_type="application/json")

class PreguntaObjAjax(TemplateView):
	def get(self, request, *args, **kwargs):
		id_pregunta = request.GET['id_pregunta']
		preguntas = Pregunta.objects.filter(id=id_pregunta)
		contenido = Contenido.objects.filter(id__in=preguntas.values('contenido'))
		data = serializers.serialize('json', contenido)
		return HttpResponse(data, content_type="application/json")
		
def busqueda(request):
	pregunta = Pregunta.objects.filter(status=True, tipo_pregunta__startswith=request.GET['consulta']).values('tipo_pregunta')
	return HttpResponse(json.dumps(list(pregunta)), content_type='application/json')

class PreguntaBusquedaAjax(TemplateView):
	def get(self, request, *args, **kwargs):
		tipo_respuesta = request.GET['tipo_respuesta']
		dificultad_filtro = request.GET['dificultad_filtro']
		contenido_filtro = request.GET['contenido_filtro']
		get_palabras = request.GET['get_palabras']
		asignatura = request.GET['asignatura']
		asignatura = int(asignatura)
		preguntas = Pregunta.objects.filter(asignatura=asignatura, tipo_pregunta__icontains=tipo_respuesta, 
			dificultad__icontains=dificultad_filtro, pregunta__icontains=get_palabras, status=True)
		data = serializers.serialize('json', preguntas)
		return HttpResponse(data, content_type="application/json")
	
# class AsignaturaAjax(TemplateView):
# 	def get(self, request, *args, **kwargs):
# 		tipo_programa = request.GET['tipo_programa']
# 		asignatura = Asignatura.objects.filter(tipo_programa=tipo_programa)
# 		data = serializers.serialize('json', asignatura)
# 		return HttpResponse(data, content_type="application/json")

# class ContenidoAjax(TemplateView):
# 	def get(self, request, *args, **kwargs):
# 		asignatura = request.GET['asignatura']
# 		# contenido = Contenido.objects.filter(codigo__codigo=asignatura)
# 		contenido = Unidad.objects.filter(asignatura__id=asignatura)
# 		# id = serializers.Field()
# 		data = serializers.serialize('json', contenido)
# 		return HttpResponse(data, content_type="application/json")

# class ObjAprendizajeAjax(TemplateView):
# 	def get(self, request, *args, **kwargs):
# 		content = request.GET['content']
# 		aprendizaje = Sesion.objects.filter(unidad_id=content)
# 		# id = serializers.Field()
# 		data = serializers.serialize('json', aprendizaje)
# 		return HttpResponse(data, content_type="application/json")












# def evaluacionPDF(request, evaluacion_id):
# 	evaluacion = Evaluacion.objects.get(id=evaluacion_id)

# 	#Inicializar pdf
# 	nombre_archivo = 'Evaluacion_%s_%s.pdf' % (evaluacion.seccion.nrc ,evaluacion.seccion.asignatura.codigo)
# 	response = HttpResponse(content_type='application/pdf')
# 	contenido = "attachment; filename={0}".format(nombre_archivo)
# 	response["Content-Disposition"] = contenido

# 	#crear el buffer
# 	buffer = BytesIO()
# 	c = canvas.Canvas(buffer, pagesize=letter)

# 	#header
# 	c.setLineWidth(.3)
# 	c.setFont('Helvetica', 22)
# 	c.drawString(30,750, 'Universidad Andrés Bello')




# 	#cerrar y guardar pdf
# 	c.save()
# 	pdf = buffer.getvalue()
# 	buffer.close()
# 	response.write(pdf)
# 	return response

		
