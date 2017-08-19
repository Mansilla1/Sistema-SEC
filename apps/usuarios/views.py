#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Código desarrollado por Daniel Mansilla
from django.shortcuts import render, render_to_response, redirect
from django.views.generic import ListView, CreateView, DetailView, DeleteView, UpdateView, TemplateView
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseBadRequest
from django.core.urlresolvers import reverse_lazy
from django.core import serializers


from django.contrib.auth.models import User
from django.contrib.auth import login, logout
from django.utils import timezone

from django.template import RequestContext
# import django_excel as excel
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO

from Proyecto.utilities import generate_pdf

from random import randrange, uniform
import random

import requests
import ast
from django.db.models import Count
# from django.utils import simplejson
# import simplejson

from .forms import *
from .models import *


def login_user(request):
    template_name = 'login/login.html'
    logout(request)
    username = password = ''
    request.session['token'] = None
    if request.POST:
        post_data = {'username': request.POST["username"],'password':request.POST["password"]}
        response = requests.post('http://cubesoa.asuscomm.com:90/rest-auth/login/', data=post_data)
        content = response.content
        content = ast.literal_eval(content)
        if "key" in content:
            post_data2 = {'username': str(request.POST["username"])}
            header = {'Content-Type':'application/json','Authorization':'Token ' + content['key']}
            response2 = requests.get('http://cubesoa.asuscomm.com:90/rest-auth/user/',headers=header, data=post_data2)
            content2 = response2.content
            content2 = ast.literal_eval(content2)
            request.session["pk"] = content2['pk']
            request.session["first_name"] = content2['first_name']
            request.session["last_name"] = content2['last_name']
            request.session["email"] = content2['email']
            request.session["token"] = content['key']
            return HttpResponseRedirect(reverse_lazy('inicio'))
        # elif 
            # return redirect('inicio')
    return render(request, template_name, {})


# def index(request):
# 	# perfil_user = PerfilUsuario.objects.get(user__id=request.user.id)
# 	# usuarios = PerfilUsuario.objects.all().count()
# 	# contenidos = ProfeSesion.objects.filter(status=True).count()
# 	# preguntas = Pregunta.objects.filter(status=True).count()
# 	# evaluaciones = Evaluacion.objects.filter(disponible=True).count()
# 	# usuario_registrados = PerfilUsuario.objects.all().order_by('created_at')[:5].reverse()

# 	# ------------------------------------------
# 	#	OBTENER RANKINGS
# 	# ------------------------------------------
# 	user_pregunta = User.objects.exclude(perfilusuario__tipo_usuario='Estudiante').annotate(preguntas=Count('pregunta')).order_by('-preguntas')[:5]
# 	user_evaluacion = User.objects.exclude(perfilusuario__tipo_usuario='Estudiante').annotate(evaluaciones=Count('evaluacion')).order_by('-evaluaciones')[:5]

# 	orden_preguntas = Pregunta.objects.all().order_by('-cant_usada')[:5]

# 	context = {
# 		'user_pregunta': user_pregunta,
# 		'user_evaluacion': user_evaluacion,
# 		'orden_preguntas': orden_preguntas,
# 		# 'perfil_user': perfil_user,
# 		# 'usuarios': usuarios,
# 		# 'preguntas': preguntas,
# 		# 'contenidos': contenidos,
# 		# 'evaluaciones': evaluaciones,
# 		# 'usuario_registrados': usuario_registrados,
# 	}
# 	return render(request, 'index.html', context)

def usuarioList(request):
	usuario = PerfilUsuario.objects.all()

	
	context = {
		'usuario': usuario,
	}

	return render(request, 'apps/usuarios/usuario_list.html', context)

def usuarioCreate(request):
	if request.POST:
		form = PerfilForm(request.POST)
		form2 = RegistroForm(request.POST)
		if form.is_valid() and form2.is_valid():
			form2 = form2.save(commit=False)
			form2.save()
			form = form.save(commit=False)
			form.user = form2
			form.save()

			#Obtener el nombre de usuario
			user=form2.username
			nombre = form.nombres + ' ' +form.apellido1 + ' ' +form.apellido2
			contrasena = 'unab2020'
			correo = form.email 
			tipouser = form.tipo_usuario
			subject = 'Bienvenido al Portal SEC!'
			message = 'Hola %s!\nusuario: %s, password: %s' % (nombre, user, contrasena)

			send_mail(
			    subject,
			    message,
			    settings.EMAIL_HOST_USER,
			    [correo],
			    fail_silently=False,
			)

		return redirect('usuarios:listar')

	else:
		form = PerfilForm()
		form2 = RegistroForm()

	context = {
		'form': form,
		'form2': form2,
	}

	return render(request, 'apps/usuarios/usuario_create.html', context)


def usuarioUpdate(request, usuario_id):
	# usuario = User.objects.get(id=usuario_id)
	# id_user = int(usuario.id)
	perfil = PerfilUsuario.objects.get(user=usuario_id)

	if request.method == 'GET':
		# form = RegistroForm(instance=usuario)
		form = PerfilForm(instance=perfil)
	else:
		# form = RegistroForm(request.POST, instance=usuario)
		form = PerfilForm(request.POST, instance=perfil)

		if form.is_valid():
			form.save()
			# form2 = form2.save(commit=False)
			# form2.user = usuario
			# form2.save()
		return redirect('usuarios:listar')

	context = {
		'form': form,
		# 'form2': form2,
		'perfil': perfil,
	}
	
	return render(request, 'apps/usuarios/usuario_update.html', context)

# class UsuarioDetail(DetailView):
# 	model = PerfilUsuario
# 	template_name = 'apps/usuarios/usuario_detail.html'
# 	context_object_name = 'usuario'

# 	def get_context_data(self, **kwargs):
# 		context = super(UsuarioDetail, self).get_context_data(**kwargs)
# 		context['title'] = 'Detalle de usuario'
# 		return context

def usuarioDelete(request, usuario_id):
	usuario = User.objects.get(id=usuario_id)
	if request.method == 'POST':
		usuario.delete()
		return redirect('usuarios:listar')
	return render(request, 'apps/usuarios/usuario_delete.html', {'usuario':usuario})

#Planillas excel
def get_planilla_usuario(request):
	#generar excel
	wb = Workbook()
	# ws = wb.create_sheet("Calificaciones",0)
	ws = wb.active
	ws.title = 'Usuarios'
	# ws.font = ws.font.copy(bold=True, italic=True)
	# Cabeceras
	a1 = ws.cell(row=1, column=1, value='RUN')
	a2 = ws.cell(row=1, column=2, value='Nombres')
	a3 = ws.cell(row=1, column=3, value='Apellido Paterno')
	a4 = ws.cell(row=1, column=4, value='Apellido Materno')
	a5 = ws.cell(row=1, column=5, value='Email')
	a6 = ws.cell(row=1, column=6, value='Usuario')
	a7 = ws.cell(row=1, column=7, value='Tipo Usuario')
	# a7 = ws.cell(row=1, column=7, value='¿Coordinador de asignatura? (si/no)')

	a1.font = Font(bold=True)
	a2.font = Font(bold=True)
	a3.font = Font(bold=True)
	a4.font = Font(bold=True)
	a5.font = Font(bold=True)
	a6.font = Font(bold=True)
	a7.font = Font(bold=True)


	nombre_archivo = 'Planilla_usuarios.xlsx'
	response = HttpResponse(content_type="application/ms-excel")
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response



def upload(request):
	if request.POST:
		# iniciar excel
		excel = request.FILES['archivo'].read() # capturar archivo
		wb = openpyxl.load_workbook(filename=BytesIO(excel)) # iniciar archivo
		hojas = wb.get_sheet_names() # capturar nombre de hojas del archivo
		hoja = wb.get_sheet_by_name(hojas[0]) #utilizar la primera hoja del documento

		total_filas = hoja.max_row # capturar valor maximo de filas a leer
		total_columnas = hoja.max_column # capturar valor maximo de columnas

		user_no_register = []
		#loop de lectura y escritura
		for i in range(2, total_filas+1):
			form = PerfilForm() #formulario de perfil usuario
			form2 = RegistroForm(request.POST or None) # formulario registro usuario
			# form3 = ProfesorForm()
			# form3_2 = CoordinadorForm()
			# form4 = EstudianteForm()

			#probar existencia de usuario
			username = hoja.cell(row=i,column=6).value
			print username
			try:
				usuario = User.objects.get(username=username)
				usuario = usuario.username
				print 'usuario ya existe'
				user_no_register += [usuario]

			except:
				rut = hoja.cell(row=i,column=1).value
				nombre = hoja.cell(row=i,column=2).value
				apellido1 = hoja.cell(row=i,column=3).value
				apellido2 = hoja.cell(row=i,column=4).value
				correo = hoja.cell(row=i,column=5).value
				usuario = hoja.cell(row=i,column=6).value
				tipo_usuario = hoja.cell(row=i,column=7).value

				nombre = nombre.capitalize()
				apellido1 = apellido1.capitalize()
				apellido2 = apellido2.capitalize()
				tipo_usuario = tipo_usuario.capitalize()

				if tipo_usuario == 'Comité académico' or tipo_usuario == 'Comite académico' or tipo_usuario == 'Comité academico':
					tipo_usuario = 'Comite academico'
				
				print tipo_usuario


				# numero_random = randrange(100,999)

				# contrasena = "%s%s%s%s" % (nombre[0].capitalize(),numero_random, apellido[:2], numero_random)
				contrasena = "unab2020"

				# form2.set_password(self.cleaned_data["password1"])
				# form2.set_password(self.cleaned_data["password2"])

				form2 = form2.save(commit=False)
				form2.username = usuario
				# form2.first_name = nombre
				# form2.last_name = apellido
				# form2.email = correo
				form2.password1 = contrasena
				form2.password2 = contrasena
				form2.save()

				form = form.save(commit=False)
				form.rut = rut
				form.nombres = nombre
				form.apellido1 = apellido1
				form.apellido2 = apellido2
				form.email = correo
				form.tipo_usuario = tipo_usuario
				form.user = form2 
				form.save()

				# if form.tipo_usuario == 'Docente':
				# 	form3 = form3.save(commit=False)
				# 	form3.usuario = form
				# 	form3.save()

				# 	# if coordinador=='si' or coordinador=='SI' or coordinador=='Si' or coordinador=='sI':
				# 	# 	form3_2 = form3_2.save(commit=False)
				# 	# 	form3_2.profesor = form
				# 	# 	form3_2.save()

				# elif form.tipo_usuario == 'Estudiante':
				# 	form4 = form4.save(commit=False)
				# 	form4.usuario = form
				# 	form4.save()

				#Obtener el nombre de usuario
				user  =form2.username
				nombre = "%s %s %s" %(form.nombres, form.apellido1, form.apellido2)
				correo = form.email 
				tipouser = form.tipo_usuario
				subject = 'Bienvenido al Portal SEC!'
				message = 'usuario: %s, password %s' % (user, contrasena)

				send_mail(
				    subject,
				    message,
				    settings.EMAIL_HOST_USER,
				    [correo],
				    fail_silently=False,
				)

		print user_no_register

		return redirect('usuarios:listar')

	else:
		form = PerfilForm()
		form2 = RegistroForm()
		# form3 = ProfesorForm()
		# # form3_2 = CoordinadorForm()
		# form4 = EstudianteForm()

	context = {
		'form': form,
		'form2': form2,
		# 'form3': form3,
		# # 'form3_2': form3_2,
		# 'form4': form4,
	}

	return render(request, 'apps/usuarios/usuario_upload.html', context)


#	ESTUDIANTES
def estudiante_list(request):
	estudiantes = Estudiante.objects.all()
	return render(request, 'apps/usuarios/estudiantes_list.html', {'estudiantes': estudiantes})

def estudiante_create(request):
	if request.POST:
		form = EstudianteForm(request.POST)
		if form.is_valid():
			form.save()
			return redirect('usuarios:listar_estudiantes')
	else:
		form = EstudianteForm()

	context = {'form': form}
	return render(request, 'apps/usuarios/estudiante_create.html', context)


def upload_estudiante(request):
	if request.POST:
		# iniciar excel
		excel = request.FILES['archivo'].read() # capturar archivo
		wb = openpyxl.load_workbook(filename=BytesIO(excel)) # iniciar archivo
		hojas = wb.get_sheet_names() # capturar nombre de hojas del archivo
		hoja = wb.get_sheet_by_name(hojas[0]) #utilizar la primera hoja del documento

		total_filas = hoja.max_row # capturar valor maximo de filas a leer
		total_columnas = hoja.max_column # capturar valor maximo de columnas

		user_no_register = []
		#loop de lectura y escritura
		for i in range(2, total_filas+1):
			form = EstudianteForm() #formulario de perfil usuario

			#	probar existencia de estudiante
			rut = hoja.cell(row=i,column=1).value
			estudiante_no_register = []
			try:
				estudiante = Estudiante.objects.get(rut=rut)
				print 'estudiante ya existe'
				estudiante_no_register += [estudiante]

			except:
				rut = hoja.cell(row=i,column=1).value
				nombres = hoja.cell(row=i,column=2).value
				apellido1 = hoja.cell(row=i,column=3).value
				apellido2 = hoja.cell(row=i,column=4).value
				correo = hoja.cell(row=i,column=5).value

				nombre2 = ''
				nombre3 = ''

				nombres = nombres.capitalize()
				apellido1 = apellido1.capitalize()
				apellido2 = apellido2.capitalize()

				form = form.save(commit=False)
				form.rut = rut
				form.nombre1 = nombres
				form.nombre2 = nombre2
				form.nombre3 = nombre3
				form.apellido1 = apellido1
				form.apellido2 = apellido2
				form.email = correo
				form.save()


		print estudiante_no_register

		return redirect('usuarios:listar_estudiantes')

	else:
		form = EstudianteForm()

	context = {
		'form': form,
	}

	return render(request, 'apps/usuarios/estudiante_upload.html', context)


def get_planilla_estudiante(request):
	#generar excel
	wb = Workbook()
	# ws = wb.create_sheet("Calificaciones",0)
	ws = wb.active
	ws.title = 'Estudiantes'
	# ws.font = ws.font.copy(bold=True, italic=True)
	# Cabeceras
	a1 = ws.cell(row=1, column=1, value='RUN')
	a2 = ws.cell(row=1, column=2, value='Nombres')
	a3 = ws.cell(row=1, column=3, value='Apellido Paterno')
	a4 = ws.cell(row=1, column=4, value='Apellido Materno')
	a5 = ws.cell(row=1, column=5, value='Email')

	a1.font = Font(bold=True)
	a2.font = Font(bold=True)
	a3.font = Font(bold=True)
	a4.font = Font(bold=True)
	a5.font = Font(bold=True)


	nombre_archivo = 'Planilla_estudiantes.xlsx'
	response = HttpResponse(content_type="application/ms-excel")
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response



#	AJAX
class GetEstudiantes(TemplateView):
	def get(self, request, *args, **kwargs):
		estudiante = Estudiante.objects.all()
		print estudiante
		data = serializers.serialize('json', estudiante)
		return HttpResponse(data, content_type="application/json")